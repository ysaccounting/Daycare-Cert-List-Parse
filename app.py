import os
import io
import json
import base64
import re
import traceback
import threading
import uuid
import anthropic
from flask import Flask, request, jsonify, send_file, Response, stream_with_context
from pdf2image import convert_from_bytes
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024 * 10  # 500MB (multiple files)

ALLOWED_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png'}

api_key = os.environ.get("ANTHROPIC_API_KEY")
if not api_key:
    raise RuntimeError("ANTHROPIC_API_KEY environment variable is not set")
client = anthropic.Anthropic(api_key=api_key)

# ── Inline HTML ────────────────────────────────────────────────────────────────

INDEX_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>DHS Child Care Certificate Parser</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600&display=swap');
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'DM Sans', sans-serif; background: #f4f6fa; min-height: 100vh; color: #1a2a3a; }
  .header { background: linear-gradient(135deg, #1a3a5c 0%, #2e6da4 100%); padding: 2rem 2.5rem; color: white; }
  .header h1 { font-family: 'DM Serif Display', serif; font-size: 1.9rem; margin-bottom: 0.3rem; }
  .header p { color: #b8d4f0; font-size: 0.95rem; }
  .container { max-width: 960px; margin: 2rem auto; padding: 0 1.5rem; }
  .card { background: white; border-radius: 12px; padding: 2rem; box-shadow: 0 2px 12px rgba(0,0,0,0.07); margin-bottom: 1.5rem; }

  /* Upload zone */
  .upload-zone { border: 2px dashed #adc8e8; border-radius: 10px; padding: 2.5rem 2rem; text-align: center; cursor: pointer; transition: all 0.2s; background: #f8fbff; }
  .upload-zone:hover, .upload-zone.drag-over { border-color: #2e6da4; background: #eef5ff; }
  .upload-zone .icon { font-size: 2.8rem; margin-bottom: 0.6rem; }
  .upload-zone p { color: #4a6a8a; font-size: 0.95rem; }
  #file-input { display: none; }

  /* File queue */
  .file-queue { margin-top: 1rem; display: none; }
  .file-item { display: flex; align-items: center; justify-content: space-between; padding: 0.5rem 0.75rem; background: #f0f6ff; border-radius: 8px; margin-bottom: 0.4rem; font-size: 0.88rem; }
  .file-item .file-name { font-weight: 500; color: #1a3a5c; flex: 1; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; margin-right: 1rem; }
  .file-item .file-status { font-size: 0.78rem; color: #7a9ab8; white-space: nowrap; }
  .file-item.processing { background: #e8f4ff; }
  .file-item.done { background: #e8f8ee; }
  .file-item.error { background: #fff0f0; }
  .remove-file { background: none; border: none; cursor: pointer; color: #aaa; font-size: 1rem; padding: 0 0.25rem; line-height: 1; }
  .remove-file:hover { color: #c00; }

  /* Buttons */
  .btn-row { display: flex; gap: 0.75rem; margin-top: 1.25rem; }
  .btn { display: inline-flex; align-items: center; justify-content: center; gap: 0.5rem; padding: 0.75rem 1.5rem; border: none; border-radius: 8px; font-family: inherit; font-size: 0.95rem; font-weight: 600; cursor: pointer; transition: all 0.2s; flex: 1; }
  .btn-primary { background: #2e6da4; color: white; }
  .btn-primary:hover { background: #1a3a5c; }
  .btn-primary:disabled { background: #9ab8d4; cursor: not-allowed; }
  .btn-danger { background: #fff; color: #c00; border: 1.5px solid #ffcccc; flex: 0 0 auto; padding: 0.75rem 1.25rem; }
  .btn-danger:hover { background: #fff0f0; }
  .btn-danger:disabled { opacity: 0.4; cursor: not-allowed; }
  .btn-success { background: #1a7a4a; color: white; }
  .btn-success:hover { background: #145c38; }
  .btn-success:disabled { background: #8ab8a0; cursor: not-allowed; }
  .btn-lg { padding: 1rem 2rem; font-size: 1.05rem; }

  /* Progress */
  .progress-wrap { margin-top: 1.25rem; display: none; }
  .progress-bar-outer { height: 10px; background: #dde8f5; border-radius: 99px; overflow: hidden; margin-bottom: 0.5rem; }
  .progress-bar-inner { height: 100%; background: #2e6da4; border-radius: 99px; width: 0%; transition: width 0.3s ease; }
  .progress-text { font-size: 0.88rem; color: #4a6a8a; margin-bottom: 0.3rem; }
  .progress-log { font-size: 0.78rem; color: #7a9ab8; max-height: 100px; overflow-y: auto; font-family: monospace; background: #f4f8ff; border-radius: 6px; padding: 0.4rem 0.6rem; display: none; }

  .error-box { background: #fff0f0; border: 1px solid #ffcccc; border-radius: 8px; padding: 1rem; color: #c00; font-size: 0.88rem; margin-top: 1rem; display: none; }

  /* Results */
  #results-section { display: none; }
  .results-header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 1rem; }
  .section-title { font-family: 'DM Serif Display', serif; font-size: 1.25rem; color: #1a3a5c; margin: 0; }
  .info-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 1rem; margin-bottom: 1.5rem; }
  .info-card { background: #f0f6ff; border-left: 4px solid #2e6da4; border-radius: 0 8px 8px 0; padding: 0.85rem 1rem; }
  .info-card h4 { font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.05em; color: #4a6a8a; margin-bottom: 0.2rem; }
  .info-card p { font-size: 0.95rem; font-weight: 600; color: #1a3a5c; }
  .metrics { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; margin-bottom: 1.5rem; }
  .metric { background: #1a3a5c; color: white; border-radius: 10px; padding: 1rem; text-align: center; }
  .metric-value { font-size: 1.8rem; font-weight: 700; }
  .metric-label { font-size: 0.78rem; opacity: 0.75; margin-top: 0.2rem; }
  .table-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
  th { background: #2e6da4; color: white; padding: 0.6rem 0.75rem; text-align: left; font-weight: 600; white-space: nowrap; }
  td { padding: 0.5rem 0.75rem; border-bottom: 1px solid #e8eef5; }
  tr:nth-child(even) td { background: #f0f6ff; }
  tr:hover td { background: #ddeeff; }
  td.num { text-align: right; font-variant-numeric: tabular-nums; }
  .editable:focus { background: #fffde7 !important; outline: none; }
  .hint { font-size: 0.8rem; color: #7a9ab8; margin-top: 0.4rem; }

  td.source-badge { font-size: 0.72rem; color: #7a9ab8; }
</style>
</head>
<body>
<div class="header">
  <h1>&#129417; DHS Child Care Certificate Parser</h1>
  <p>Upload one or more DHS Child Care Certificate Reports &mdash; Claude reads each page and generates formatted Excel files.</p>
</div>
<div class="container">
  <div class="card">
    <div class="upload-zone" id="upload-zone">
      <div class="icon">&#128196;</div>
      <p><strong>Drop files here</strong> or click to browse</p>
      <p style="margin-top:0.4rem;font-size:0.82rem;">PDF, JPG, or PNG &middot; Multiple files supported &middot; up to 50 MB each</p>
    </div>
    <input type="file" id="file-input" accept=".pdf,.jpg,.jpeg,.png" multiple>

    <div class="file-queue" id="file-queue"></div>

    <div class="progress-wrap" id="progress-wrap">
      <div class="progress-text" id="progress-text"></div>
      <div class="progress-bar-outer"><div class="progress-bar-inner" id="progress-bar"></div></div>
      <div class="progress-log" id="progress-log"></div>
    </div>

    <div class="error-box" id="error-box"></div>

    <div class="btn-row">
      <button class="btn btn-primary btn-lg" id="parse-btn" disabled>&#10024; Parse Certificates</button>
      <button class="btn btn-danger" id="clear-btn" disabled title="Clear all files and results">&#10006; Clear</button>
    </div>
  </div>

  <div id="results-section">
    <div class="card">
      <div class="results-header">
        <div class="section-title" id="children-title">Children</div>
        <button class="btn btn-success" id="download-btn">&#11015;&#65039; Download Excel</button>
      </div>
      <div class="info-grid" id="info-grid"></div>
      <div class="metrics" id="metrics"></div>
      <p class="hint" style="margin-bottom:0.75rem;">Click any cell to edit before downloading.</p>
      <div class="table-wrap">
        <table>
          <thead><tr>
            <th>Parent/Client</th><th>Child's Name</th>
            <th>Rate ($)</th><th>Co-Pay ($)</th><th>Elig Days</th>
            <th>Attd Days</th><th>Adj Days</th><th>Payment ($)</th>
          </tr></thead>
          <tbody id="table-body"></tbody>
        </table>
      </div>
      <div style="margin-top:1.5rem;">
        <button class="btn btn-success btn-lg" id="download-btn-2">&#11015;&#65039; Download Excel</button>
      </div>
    </div>
  </div>
</div>

<script>
// ── State ──────────────────────────────────────────────────────────────────
let fileQueue = [];
let parsedData = null;
let isRunning  = false;

const uploadZone   = document.getElementById('upload-zone');
const fileInput    = document.getElementById('file-input');
const fileQueueEl  = document.getElementById('file-queue');
const parseBtn     = document.getElementById('parse-btn');
const clearBtn     = document.getElementById('clear-btn');
const progressWrap = document.getElementById('progress-wrap');
const progressBar  = document.getElementById('progress-bar');
const progressText = document.getElementById('progress-text');
const progressLog  = document.getElementById('progress-log');
const errorBox     = document.getElementById('error-box');
const resultsSection = document.getElementById('results-section');
const downloadBtn  = document.getElementById('download-btn');
const downloadBtn2 = document.getElementById('download-btn-2');

// ── File handling ──────────────────────────────────────────────────────────
const ALLOWED = ['.pdf','.jpg','.jpeg','.png'];
function isAllowed(name) { return ALLOWED.some(e => name.toLowerCase().endsWith(e)); }

function addFiles(files) {
  for (const file of files) {
    if (!isAllowed(file.name)) continue;
    if (fileQueue.some(f => f.file.name === file.name && f.file.size === file.size)) continue;
    fileQueue.push({ file, id: crypto.randomUUID(), status: 'pending' });
  }
  renderQueue();
  parseBtn.disabled = fileQueue.length === 0 || isRunning;
  clearBtn.disabled = fileQueue.length === 0 && !parsedData;
}

function removeFile(id) {
  fileQueue = fileQueue.filter(f => f.id !== id);
  renderQueue();
  parseBtn.disabled = fileQueue.length === 0 || isRunning;
  clearBtn.disabled = fileQueue.length === 0 && !parsedData;
}

function renderQueue() {
  if (fileQueue.length === 0) { fileQueueEl.style.display = 'none'; return; }
  fileQueueEl.style.display = 'block';
  fileQueueEl.innerHTML = fileQueue.map(f => {
    const icons = { pending: '&#9711;', processing: '&#9680;', done: '&#9989;', error: '&#10060;' };
    const statusText = { pending: 'queued', processing: 'reading\u2026', done: 'done', error: 'error' };
    const cls = f.status !== 'pending' ? f.status : '';
    return `<div class="file-item ${cls}" id="fitem-${f.id}">
      <span class="file-name">${icons[f.status]} ${f.file.name}</span>
      <span class="file-status">${statusText[f.status]}</span>
      ${f.status === 'pending' ? `<button class="remove-file" onclick="removeFile('${f.id}')" title="Remove">&#10005;</button>` : ''}
    </div>`;
  }).join('');
}

function setFileStatus(id, status) {
  const f = fileQueue.find(f => f.id === id);
  if (f) { f.status = status; renderQueue(); }
}

// ── Upload zone events ─────────────────────────────────────────────────────
uploadZone.addEventListener('click', () => fileInput.click());
fileInput.addEventListener('change', () => { addFiles(fileInput.files); fileInput.value = ''; });
uploadZone.addEventListener('dragover', e => { e.preventDefault(); uploadZone.classList.add('drag-over'); });
uploadZone.addEventListener('dragleave', () => uploadZone.classList.remove('drag-over'));
uploadZone.addEventListener('drop', e => {
  e.preventDefault(); uploadZone.classList.remove('drag-over');
  addFiles(e.dataTransfer.files);
});

// ── Clear ──────────────────────────────────────────────────────────────────
clearBtn.addEventListener('click', () => {
  if (isRunning) return;
  fileQueue = [];
  parsedData = null;
  renderQueue();
  progressWrap.style.display = 'none';
  progressLog.innerHTML = '';
  progressLog.style.display = 'none';
  progressBar.style.width = '0%';
  errorBox.style.display = 'none';
  resultsSection.style.display = 'none';
  document.getElementById('table-body').innerHTML = '';
  parseBtn.disabled = true;
  clearBtn.disabled = true;
  fileInput.value = '';
});

// ── Parse ──────────────────────────────────────────────────────────────────
function addLog(msg) {
  progressLog.style.display = 'block';
  const line = document.createElement('div');
  line.textContent = msg;
  progressLog.appendChild(line);
  progressLog.scrollTop = progressLog.scrollHeight;
}

async function processFile(entry, fileIndex, totalFiles) {
  const { file, id } = entry;
  setFileStatus(id, 'processing');

  const form = new FormData();
  form.append('file', file);
  let jobId;
  try {
    const resp = await fetch('/upload', { method: 'POST', body: form });
    const data = await resp.json();
    if (!resp.ok || data.error) throw new Error(data.error || 'Upload failed');
    jobId = data.job_id;
  } catch (err) {
    setFileStatus(id, 'error');
    addLog(`\u274C ${file.name}: ${err.message}`);
    return null;
  }

  return new Promise(resolve => {
    const evtSource = new EventSource('/progress/' + jobId);
    evtSource.addEventListener('progress', e => {
      const d = JSON.parse(e.data);
      const filePct   = d.pct;
      const globalPct = Math.round(((fileIndex / totalFiles) + (filePct / 100 / totalFiles)) * 100);
      progressBar.style.width = globalPct + '%';
      progressText.textContent = `File ${fileIndex+1}/${totalFiles}: ${file.name} \u2014 ${d.msg}`;
      if (d.detail) addLog(`[${file.name}] ${d.detail}`);
    });
    evtSource.addEventListener('done', e => {
      evtSource.close();
      const d = JSON.parse(e.data);
      if (d.error) {
        setFileStatus(id, 'error');
        addLog(`\u274C ${file.name}: ${d.error}`);
        resolve(null);
      } else {
        setFileStatus(id, 'done');
        addLog(`\u2705 ${file.name}: ${d.count} rows extracted`);
        (d.children || []).forEach(c => c._source = file.name);
        resolve(d);
      }
    });
    evtSource.onerror = () => {
      evtSource.close();
      setFileStatus(id, 'error');
      addLog(`\u274C ${file.name}: connection lost`);
      resolve(null);
    };
  });
}

parseBtn.addEventListener('click', async () => {
  if (fileQueue.length === 0 || isRunning) return;
  isRunning = true;
  parseBtn.disabled = true;
  clearBtn.disabled = true;
  errorBox.style.display = 'none';
  resultsSection.style.display = 'none';
  progressWrap.style.display = 'block';
  progressLog.innerHTML = '';
  progressLog.style.display = 'none';
  progressBar.style.width = '2%';
  progressText.textContent = 'Starting\u2026';

  const results = [];
  const pendingFiles = fileQueue.filter(f => f.status !== 'done');
  pendingFiles.forEach(f => { f.status = 'pending'; });
  renderQueue();

  for (let i = 0; i < pendingFiles.length; i++) {
    const result = await processFile(pendingFiles[i], i, pendingFiles.length);
    if (result) results.push(result);
  }

  isRunning = false;
  progressBar.style.width = '100%';

  if (results.length === 0) {
    progressText.textContent = 'No records extracted.';
    errorBox.textContent = '\u26A0\uFE0F No records could be extracted. Check the log for details.';
    errorBox.style.display = 'block';
    parseBtn.disabled = false;
    clearBtn.disabled = false;
    return;
  }

  const merged = mergeResults(results);
  parsedData = merged;
  progressText.textContent = `Done \u2014 ${merged.children.length} rows from ${results.length} file(s).`;
  renderResults(merged);
  resultsSection.style.display = 'block';
  parseBtn.disabled = false;
  clearBtn.disabled = false;
});

// ── Merge results from multiple files ─────────────────────────────────────
function mergeResults(results) {
  const provider = results[0]?.provider || {};
  const summary  = results[results.length - 1]?.summary || {};
  const children = results.flatMap(r => r.children || []);
  return { provider, children, summary, count: children.length };
}

// ── Render ─────────────────────────────────────────────────────────────────
// Payment priority: adj_days > attd_days > elig_days
function calcPayment(rate, copay, elig, attd, adj) {
  rate  = parseFloat(rate)  || 0;
  copay = parseFloat(copay) || 0;
  const days = (adj  !== null && adj  !== '' && adj  !== undefined) ? parseFloat(adj)
             : (attd !== null && attd !== '' && attd !== undefined) ? parseFloat(attd)
             : (parseFloat(elig) || 0);
  return (rate * days - copay).toFixed(2);
}

function renderResults(data) {
  const pi       = data.provider  || {};
  const children = data.children  || [];

  document.getElementById('info-grid').innerHTML = [
    ['Provider',       pi.name          || '\u2014'],
    ['Report Period',  pi.period        || '\u2014'],
    ['DHS Provider ID',pi.dhs_id        || '\u2014'],
    ['Date of Issue',  pi.date_of_issue || '\u2014'],
  ].map(([l,v]) => `<div class="info-card"><h4>${l}</h4><p>${v}</p></div>`).join('');

  const totalAttd = children.reduce((s,c) => s + (parseInt(c.attd_days)||0), 0);
  const totalPay  = children.reduce((s,c) => s + parseFloat(calcPayment(c.rate,c.copay,c.elig_days,c.attd_days,c.adj_days)), 0);
  document.getElementById('metrics').innerHTML = `
    <div class="metric"><div class="metric-value">${children.length}</div><div class="metric-label">Rows</div></div>
    <div class="metric"><div class="metric-value">${totalAttd}</div><div class="metric-label">Attended Days</div></div>
    <div class="metric"><div class="metric-value">$${totalPay.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2})}</div><div class="metric-label">Total Payment</div></div>`;

  document.getElementById('children-title').textContent = `Records (${children.length} rows)`;

  const tbody = document.getElementById('table-body');
  tbody.innerHTML = '';
  children.forEach((c, idx) => {
    const pay = calcPayment(c.rate, c.copay, c.elig_days, c.attd_days, c.adj_days);
    const tr = document.createElement('tr');
    tr.innerHTML = `
      <td contenteditable class="editable" data-field="parent_client">${c.parent_client||''}</td>
      <td contenteditable class="editable" data-field="child_name">${c.child_name||''}</td>
      <td contenteditable class="editable num" data-field="rate">${c.rate!=null?parseFloat(c.rate).toFixed(2):''}</td>
      <td contenteditable class="editable num" data-field="copay">${c.copay!=null?parseFloat(c.copay).toFixed(2):''}</td>
      <td contenteditable class="editable num" data-field="elig_days">${c.elig_days??''}</td>
      <td contenteditable class="editable num" data-field="attd_days">${c.attd_days??''}</td>
      <td contenteditable class="editable num" data-field="adj_days">${c.adj_days??''}</td>
      <td class="num pay-cell">$${pay}</td>`;
    tr.querySelectorAll('[contenteditable]').forEach(cell => {
      cell.addEventListener('blur', () => {
        const field = cell.dataset.field;
        let val = cell.textContent.trim();
        if (['rate','copay','elig_days','attd_days','adj_days'].includes(field))
          val = val === '' ? null : parseFloat(val);
        parsedData.children[idx][field] = val;
        const ch = parsedData.children[idx];
        tr.querySelector('.pay-cell').textContent = '$' + calcPayment(ch.rate,ch.copay,ch.elig_days,ch.attd_days,ch.adj_days);
      });
    });
    tbody.appendChild(tr);
  });
}

// ── Download ───────────────────────────────────────────────────────────────
async function doDownload() {
  if (!parsedData) return;
  downloadBtn.disabled = true; downloadBtn2.disabled = true;
  downloadBtn.textContent = '\u23F3 Building\u2026'; downloadBtn2.textContent = '\u23F3 Building\u2026';
  try {
    const resp = await fetch('/download', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(parsedData)
    });
    if (!resp.ok) throw new Error('Download failed');
    const blob = await resp.blob();
    const cd = resp.headers.get('Content-Disposition') || '';
    const match = cd.match(/filename="?([^"]+)"?/);
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a'); a.href = url; a.download = match ? match[1] : 'Certs.xlsx'; a.click();
    URL.revokeObjectURL(url);
  } catch (err) { alert('Download error: ' + err.message); }
  finally {
    downloadBtn.disabled = false; downloadBtn2.disabled = false;
    downloadBtn.textContent = '\u2B07\uFE0F Download Excel'; downloadBtn2.textContent = '\u2B07\uFE0F Download Excel';
  }
}
downloadBtn.addEventListener('click', doDownload);
downloadBtn2.addEventListener('click', doDownload);
</script>
</body>
</html>"""

# ── Job store ──────────────────────────────────────────────────────────────────
jobs = {}
jobs_lock = threading.Lock()

def push_event(job_id, event_type, data):
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id]['events'].append({'type': event_type, 'data': json.dumps(data)})

# ── Claude vision parser ───────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a data extraction assistant for Illinois DHS Child Care Certificate Reports.
When given an image of a certificate page, extract all child records and return ONLY valid JSON.
No preamble, no explanation, no markdown fences — raw JSON only.

Return this exact structure:
{
  "provider": {
    "name": "",
    "dhs_id": "",
    "period": "",
    "date_of_issue": "",
    "days_open": null
  },
  "children": [
    {
      "parent_client": "",
      "child_name": "",
      "dob": "",
      "rate": null,
      "copay": null,
      "elig_days": null,
      "attd_days": null,
      "adj_days": null
    }
  ],
  "summary": {
    "total_attended": null,
    "total_eligible": null,
    "attendance_pct": null,
    "signed_by": "",
    "sig_date": ""
  },
  "is_data_page": true
}

FORM LAYOUT — each child record block looks like this:
  12345 67890 12345          <- CHILD CARE CASE NO. (digits only, ignore this)
  IVR CASE#:3-200-123456    <- IVR line (ignore this)
  FIRSTNAME LASTNAME         <- CLIENT'S NAME — this is the PARENT. Use ONLY this name for parent_client.
    CHILD NAME               <- CHILD'S NAME — this is the child
    MM/DD/YYYY               <- child date of birth
    46.00  F  1.00  ...      <- rate rows (F, P, S)

COLUMN ORDER — count left to right on each rate row:
  Col 1: RATE              (e.g. 54.00, 27.00, 23.00)
  Col 2: TYPE              (F = full-day, P = part-day, S = school-age)
  Col 3: PARENT CO-PAY     (printed dollar amount or blank)
  Col 4: ADJUSTED CO-PAY   (printed dollar amount or blank)
  Col 5: DAYS ELIG         → elig_days  (printed number, e.g. 22)
  Col 6: ADJ ELIG          → adj_days   (handwritten — the FIRST blank/underline after DAYS ELIG)
  Col 7: ATTD              → attd_days  (handwritten — the SECOND blank/underline, closest to right margin)
  Col 8: CODE              (letter code, ignore)

CRITICAL — reading handwritten day values:
- Col 6 (ADJ ELIG) and Col 7 (ATTD) are two SEPARATE blank fields side by side.
- The handwritten number closest to the RIGHT MARGIN of the page is ALWAYS attd_days (Col 7).
- The handwritten number to the LEFT of that, between DAYS ELIG and ATTD, is adj_days (Col 6).
- If there is only ONE handwritten number on the row and it is near the right margin → attd_days=N, adj_days=null
- If there are TWO handwritten numbers → left one = adj_days, right one = attd_days
- If both blanks are empty (just blank lines ——) → both are null
- NEVER put the ATTD value (right margin) into adj_days
- NEVER put the ADJ ELIG value (middle) into attd_days
- A blank line (——) always means null, never zero

RATE ROW RULES — one output record per filled row:
- Each child has up to 3 rate rows: F (full-day), P (part-day), S (school-age).
- A row is "filled" if it has a number in the DAYS ELIG column (Col 5).
- USUALLY only one row is filled → output ONE record for that child.
- BUT if TWO OR MORE rows are filled, output ONE RECORD PER FILLED ROW,
  each with the same parent_client, child_name, and dob, but with that row's
  own rate, copay, elig_days, attd_days, and adj_days.
- Example — child with both F and P rows filled:
    {"parent_client":"OLIVIA CORONA","child_name":"MADELYN TRUJILLO","dob":"7/02/2025","rate":67.00,"copay":1.00,"elig_days":10,"attd_days":10,"adj_days":null}
    {"parent_client":"OLIVIA CORONA","child_name":"MADELYN TRUJILLO","dob":"7/02/2025","rate":34.00,"copay":0.00,"elig_days":8,"attd_days":7,"adj_days":null}
- NEVER skip a filled row. NEVER merge two filled rows into one record.
- rate: dollar amount from that specific row
- copay: PARENT CO-PAY from that same row; use 0.0 if blank

Other rules:
- parent_client: the CLIENT'S NAME — the all-caps adult name BELOW the IVR CASE# line.
  NEVER put the case number (digits), the IVR CASE# string, or any number in parent_client.
  Example: "IVR CASE#:3-200-299684" then "KIYATA PERRY" → parent_client = "KIYATA PERRY"
- child_name: the child's name, indented or on the next line after the client name, with a DOB below it
- is_data_page: false for blank pages, signature-only pages, or pages with no child records
- provider: fill from header if visible; empty string if not present
- children: only records with actual data (name + rate). Skip column header rows.
- elig_days/attd_days/adj_days: integers or null. Never strings.
- For children marked "no longer attends" (C code): adj_days = 0, attd_days = 0
- summary: only fill on the final signature/summary page; leave null elsewhere
- Do not invent data. If a field is blank or illegible, use null or empty string.
"""

def image_to_b64(pil_image):
    buf = io.BytesIO()
    pil_image.save(buf, format="JPEG", quality=85)
    return base64.standard_b64encode(buf.getvalue()).decode()

def get_pages(file_bytes, filename):
    ext = os.path.splitext(filename.lower())[1]
    if ext == '.pdf':
        return convert_from_bytes(file_bytes, dpi=150)
    elif ext in ('.jpg', '.jpeg', '.png'):
        return [Image.open(io.BytesIO(file_bytes)).convert("RGB")]
    raise ValueError(f"Unsupported file type: {ext}")

def parse_page(pil_image):
    b64 = image_to_b64(pil_image)
    message = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                {"type": "text", "text": "Extract all data from this DHS Child Care Certificate page. Return only JSON."}
            ]
        }]
    )
    raw = message.content[0].text.strip()
    raw = re.sub(r'```[\w]*\n?|```', '', raw).strip()
    return json.loads(raw)

def run_job(job_id, file_bytes, filename):
    try:
        pages = get_pages(file_bytes, filename)
        n = len(pages)
        push_event(job_id, 'progress', {'pct': 5, 'msg': f'Converting — {n} pages', 'detail': f'{n} pages found'})

        provider_info = {}
        all_children  = []
        summary       = {}

        for i, page in enumerate(pages):
            pct = 5 + int(90 * i / n)
            push_event(job_id, 'progress', {'pct': pct, 'msg': f'Page {i+1} of {n}…', 'detail': None})
            try:
                result = parse_page(page)
            except Exception as e:
                push_event(job_id, 'progress', {'pct': pct, 'msg': f'Page {i+1} of {n}…', 'detail': f'Page {i+1} error: {e}'})
                continue

            if not result.get('is_data_page', True):
                continue

            p = result.get('provider') or {}
            for key in ('name','dhs_id','period','date_of_issue','days_open'):
                if not provider_info.get(key) and p.get(key):
                    provider_info[key] = p[key]

            kids = [c for c in (result.get('children') or []) if c.get('child_name')]
            all_children.extend(kids)

            s = result.get('summary') or {}
            for key in ('total_attended','total_eligible','attendance_pct','signed_by','sig_date'):
                if s.get(key) not in (None, '', 0):
                    summary[key] = s[key]

            if kids:
                push_event(job_id, 'progress', {'pct': pct, 'msg': f'Page {i+1} of {n}…',
                    'detail': f'Page {i+1}: {len(kids)} rows (total: {len(all_children)})'})

        push_event(job_id, 'done', {
            'provider': provider_info, 'children': all_children,
            'summary': summary, 'count': len(all_children)
        })
    except Exception as e:
        push_event(job_id, 'done', {'error': str(e), 'count': 0})
    finally:
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]['done'] = True

# ── Excel builder ──────────────────────────────────────────────────────────────

def build_excel(provider_info, children, summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Children Detail"
    hf  = PatternFill("solid", start_color="1F4E79")
    shf = PatternFill("solid", start_color="2E75B6")
    af  = PatternFill("solid", start_color="D6E4F0")
    wf  = PatternFill("solid", start_color="FFFFFF")
    thin = Side(style="thin", color="AAAAAA")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    period = provider_info.get("period", "")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Child Care Certificate Report – {period}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hf
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for i, (lbl, val) in enumerate([
        ("Provider:", provider_info.get("name","")),
        ("DHS Provider ID#:", provider_info.get("dhs_id","")),
        ("Report Period:", period),
        ("Days Open:", provider_info.get("days_open","")),
        ("Date of Issue:", provider_info.get("date_of_issue","")),
    ], start=2):
        ws[f"A{i}"] = lbl; ws[f"A{i}"].font = Font(name="Arial", bold=True, size=10)
        ws[f"B{i}"] = val; ws[f"B{i}"].font = Font(name="Arial", size=10)

    hr = 8
    # Columns: A=Parent, B=Child, C=Rate, D=Copay, E=Elig Days, F=Attd Days, G=Adj Days, H=Payment
    for c, h in enumerate(["Parent/Client", "Child's Name",
                            "Rate ($)", "Co-Pay ($)", "Elig Days",
                            "Attd Days", "Adj Days", "Payment ($)"], 1):
        cell = ws.cell(hr, c, h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = shf; cell.border = bdr
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hr].height = 30

    for ri, child in enumerate(children, start=hr+1):
        fill = af if ri % 2 == 0 else wf
        for ci, val in enumerate([
            child.get("parent_client", ""),
            child.get("child_name", ""),
            child.get("rate"),
            child.get("copay"),
            child.get("elig_days"),
            child.get("attd_days"),
            child.get("adj_days"),
        ], 1):
            cell = ws.cell(ri, ci, val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = bdr
            cell.alignment = Alignment(horizontal="center")
            if ci in (3, 4):
                cell.number_format = '$#,##0.00'
        # Column 8 (H): Payment — priority: Adj Days (G) > Attd Days (F) > Elig Days (E)
        pay = ws.cell(ri, 8)
        pay.value = f'=IF(G{ri}<>"",C{ri}*G{ri},IF(F{ri}<>"",C{ri}*F{ri},C{ri}*E{ri}))-D{ri}'
        pay.font = Font(name="Arial", size=10)
        pay.fill = fill
        pay.border = bdr
        pay.alignment = Alignment(horizontal="center")
        pay.number_format = '$#,##0.00'

    tr = hr + len(children) + 1
    for c in range(1, 9):
        cell = ws.cell(tr, c)
        cell.border = bdr
        cell.fill = PatternFill("solid", start_color="BDD7EE")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    ws.cell(tr, 1, "TOTALS").fill = shf
    ws.cell(tr, 1).font = Font(bold=True, color="FFFFFF", name="Arial")
    ws.cell(tr, 5, f"=SUM(E{hr+1}:E{hr+len(children)})")
    ws.cell(tr, 6, f"=SUM(F{hr+1}:F{hr+len(children)})")
    ws.cell(tr, 8, f"=SUM(H{hr+1}:H{hr+len(children)})").number_format = '$#,##0.00'

    for i, w in enumerate([24, 20, 11, 11, 10, 10, 10, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Attendance Summary")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = f"Attendance Summary – {period}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = hf
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26
    pct = f"{summary.get('attendance_pct','')}%" if summary.get('attendance_pct') else ""
    for r, (lbl, val) in enumerate([
        ("Total Attended Days", summary.get("total_attended", "")),
        ("Total Adjusted Eligible Days", summary.get("total_eligible", "")),
        ("Attendance Percentage", pct),
        ("", ""),
        ("Signed By", summary.get("signed_by", "")),
        ("Signature Date", summary.get("sig_date", "")),
    ], 2):
        ws2[f"A{r}"] = lbl; ws2[f"A{r}"].font = Font(name="Arial", bold=True, size=11)
        ws2[f"B{r}"] = val; ws2[f"B{r}"].font = Font(name="Arial", size=11)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def make_filename(provider_info):
    raw_name = provider_info.get("name", "ChildCare")
    raw_period = provider_info.get("period", "")
    name_clean = re.sub(r'\b(DAYCARE|DAY CARE|LLC|INC|CORP|CENTER)\b', '', raw_name, flags=re.IGNORECASE).strip()
    name_clean = " ".join(name_clean.split()).title()
    parts = raw_period.split()
    period_short = f"{parts[0][:3].title()} {parts[1]}" if len(parts) == 2 else raw_period.title()
    return f"{name_clean} {period_short} Certs.xlsx"

# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return INDEX_HTML, 200, {"Content-Type": "text/html"}

@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    ext = os.path.splitext(f.filename.lower())[1]
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": f"Unsupported file type '{ext}'."}), 400
    file_bytes = f.read()
    job_id = str(uuid.uuid4())
    with jobs_lock:
        jobs[job_id] = {'events': [], 'done': False}
    threading.Thread(target=run_job, args=(job_id, file_bytes, f.filename), daemon=True).start()
    return jsonify({"job_id": job_id})

@app.route("/progress/<job_id>")
def progress(job_id):
    def generate():
        import time
        cursor = 0
        while True:
            with jobs_lock:
                job = jobs.get(job_id)
                if not job:
                    yield "event: done\ndata: {\"error\": \"Job not found\"}\n\n"
                    return
                events = job['events'][cursor:]
                is_done = job['done']

            for ev in events:
                yield f"event: {ev['type']}\ndata: {ev['data']}\n\n"
                cursor += 1
                if ev['type'] == 'done':
                    def cleanup(jid=job_id):
                        time.sleep(60)
                        with jobs_lock: jobs.pop(jid, None)
                    threading.Thread(target=cleanup, daemon=True).start()
                    return

            if is_done and not events:
                return
            time.sleep(0.4)

    return Response(stream_with_context(generate()),
                    mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

@app.route("/download", methods=["POST"])
def download():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON body"}), 400
        buf = build_excel(data.get("provider", {}), data.get("children", []), data.get("summary", {}))
        return send_file(buf, as_attachment=True,
                         download_name=make_filename(data.get("provider", {})),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route("/health")
def health():
    return jsonify({"status": "ok", "api_key_set": bool(api_key), "active_jobs": len(jobs)})

@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "File too large. Maximum 50MB per file."}), 413

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True, threaded=True)
